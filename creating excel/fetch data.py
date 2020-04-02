import openpyxl #fetch data from the excel file using python
path = "data9.xlsx" #importing module
data9_object = openpyxl.load_workbook(path) # gives the location of our file
data9_object = data9_object.active #active


get_data = data9_object.cell(row=2, column=1)  # gets a value
print(get_data.value)

print(data9_object.max_row, data9_object.max_column) #tells max rows and columns

data_sheet = data9_object.title # gets title
print(data_sheet)

# name_sheet = data9_object = "class9"
# print(name_sheet())
data9_object.title = "class 99"
data_sheet = data9_object.title
print(data_sheet)

data9_object.cell(row=3,column=1).value = "Changed"
print(data9_object.cell(row=3,column=1).value)

# implement changed to excel file
#hint it can be made to any python file in order to achive the task